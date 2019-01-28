# -*- coding: utf-8 -*-
import xlwings, xlrd, openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os, sys, json, time
import win32con, win32clipboard
from win32com.client import Dispatch
from win32com.client import *
import win32gui
import win32api
import pywin32
from ctypes import *
import traceback
from itertools import *
from operator import itemgetter
from concurrent import futures  # 并发

def get_continuous_number(_raw_list):
    '''
    # h ttps://blog.csdn.net/weixin_42555131/article/details/82020064
    解析一个int数组，如果有连续的，则用connector字符连接
    如： list = [1,2,3,4,8,9,10,12,18]
    返回['1:4', '8:10', 12, 18] 或者 [(1,4),(8,10),12,18 ]
    :param _raw_list: 原始int数组
    :param connector: 连接器，默认是冒号(供Excel里的Range调用)，也可以根据需求改成其他的
    :return: 简化了的数组
    '''
    fun = lambda x: x[1]-x[0]  # 比较后一位和前一位是否相差一个数值，比如相差1
    index_list = []
    for k, g in groupby(enumerate(_raw_list), fun):
        list_temp = [j for i, j in g]  # 连续的数字列表
        if len(list_temp) > 1:
            #scope = str(min(list_temp)) + connector + str(max(list_temp))  # 用:来连接，方便Excel调用，这种写法不好用，涉及类型转换烦死人
            scope = (min(list_temp)+97, max(list_temp)+97)  # 元祖大法好！！
        else:
            scope = list_temp[0]+97
        index_list.append(scope)

    _list = index_list.copy()  # 创建一个副本出来，以免干扰了传入的raw_list
    for i in range(len(_list)):
        if i == 0:
            _list[0] = _list[0]
        else:
            if isinstance(_list[i], tuple):
                delta = _list[i][1] - _list[i][0]
                if isinstance(_list[i-1], tuple):
                    _list[i] = (_list[i-1][1] + 1, _list[i-1][1] + 1 + delta)
                else:
                    _list[i] = (_list[i-1]+1, _list[i-1]+1 + delta)
            else:
                if isinstance(_list[i-1], tuple):
                    _list[i] = _list[i-1][1] + 1
                else:
                    _list[i] = _list[i-1]+1
    return index_list, _list

def move_item_to_end(_list, _item_value):
    # 将不存在重复项的list里的某个值，移动到最后
    for _value in _list:
        if _value == _item_value:
            _index = _list.index(_value)
            _list.append(_value)
            _list.remove(_list[_index])
    return _list


def create_file_from_file(source_dir, desitination_dir):
    '''
    创建文件
    :param source_dir: 当前文件夹
    :param _from: 从某种格式的文件，如txt
    :param _to: 创建同名的另一种格式的文件，如xlsx
    :return:
    '''
    # 判断传入的路径的两种方式。C:\Folder和C:\Folder统一处理成C:\Folder\
    if source_dir[-1:] != "\\":
        source_dir = source_dir + "\\"
    if desitination_dir[-1:] != "\\":
        desitination_dir = desitination_dir + "\\"
    _list = os.listdir(source_dir)

    for i in range(len(_list)):
        _filepath = os.path.join(source_dir, _list[i])
        # if os.path.isfile(_filepath):
        _app = Dispatch("Excel.Application")
        _workbook = _app.Workbooks.Add()
        _workbook.SaveAs(desitination_dir + _list[i].split(".txt")[0] + ".xlsx")
        _workbook.Close()


def write_txts_to_xlsxs(txtrootpath, excelrootpath):
    # 判断传入的路径的两种方式。C:\Folder和C:\Folder统一处理成C:\Folder\
    if txtrootpath[-1:] != "\\":
        txtrootpath = txtrootpath + "\\"
    if excelrootpath[-1:] != "\\":
        excelrootpath = excelrootpath + "\\"

    txt_file_list = os.listdir(txtrootpath)
    excel_file_list = os.listdir(excelrootpath)
    #not_done_index = [306, 354, 357, 358, 359, 362, 366, 311, 368, 369, 370, 312, 372, 373]  程序中途断电，找到目录下未完成的excel在所有excel 的索引，然后接着写
    #not_done_filename =  ['戴尔 Ins 15MF', '戴尔 Ins14SD-3328', '戴尔 Ins14UD-1328', '戴尔 Ins14UD-1528', '戴尔 Ins14UD-1548', '戴尔 Ins14UD-2108', '戴尔 Ins14UD-3548', '戴尔 Ins14UD-3748', '戴尔 Ins14UD-4328', '戴尔 Ins14UD-4528', '戴尔 Ins14UD-4548', '戴尔 Ins14UD-4728', '戴尔 Ins14UD-5625', '戴尔 Ins14UD-5725']
    for i in range(len(excel_file_list)):
        #if i not in not_done_index: continue
        excel_path = excelrootpath + excel_file_list[i]  # 构造excel的绝对路径
        txt_path = txtrootpath + txt_file_list[i].split(".txt")[0] + ".txt"
        sheet_name = excel_file_list[i].split(".xlsx")[0]  # 需要创建的Excel文档的Sheet1重命名为产品名

        if excel_file_list[i].split(".xlsx")[0] not in "戴尔 Ins14UD-5525": continue
        #print("1")
        # https://www.cnblogs.com/chongzi1990/p/8694883.html
        # 去除 \ufeff. 直接指定  encoding='UTF-8-sig'
        with open(txt_path, encoding='UTF-8-sig') as txt_file:
            txt_content = txt_file.read().replace("\n\n", "\r\n")

            # 下一句代码是用来处理 json.dump(_json, file_obj, indent=4, ensure_ascii=False) 生成的txt文本正文
            # txt_content_out = txt_content.replace("[", "").replace("]", "").replace("\"", "").replace("    ","").replace("\n","")

            txtlist = txt_content.split("\r\n")

            try:
                app = xlwings.App(visible=False, add_book=False)
                time.sleep(0.5)
                book = app.books.add()

                # book.sheets.add("Sheet2")  # 添加一个Sheet
                # book.sheets.__delitem__("Sheet1")  # 删除Sheet1

                for j in range(len(txtlist)):
                    xlwings.Range((j + 1, 1)).value = txtlist[j]  # xlwing从1开始
                    print("Excel文档" + sheet_name + ", 第" + str(j + 1) + "行写完了")

                book.save(excel_path)
                book.close()
                app.quit()  # 【坑】这里千万只能写quit，不能直接kill否则循环的下次就跑不动了
                time.sleep(1)  # 防止Excel压力太大，睡1秒
                print("呵呵呵")
            except:
                traceback.print_exc()
            finally:
                app.quit()
                app.kill()

def write_txts_to_xlsxs_xlwings(txtrootpath, excelrootpath):
    # 判断传入的路径的两种方式。C:\Folder和C:\Folder统一处理成C:\Folder\
    if txtrootpath[-1:] != "\\":
        txtrootpath = txtrootpath + "\\"
    if excelrootpath[-1:] != "\\":
        excelrootpath = excelrootpath + "\\"

    txt_file_list = os.listdir(txtrootpath)
    excel_file_list = os.listdir(excelrootpath)

    for i in range(len(excel_file_list)):
        if i < 131: continue  # 程序控制器
        excel_path = excelrootpath + excel_file_list[i]  # 构造excel的绝对路径
        txt_path = txtrootpath + txt_file_list[i].split(".txt")[0] + ".txt"
        sheet_name = excel_file_list[i].split(".xlsx")[0]  # 需要创建的Excel文档的Sheet1重命名为产品名

        # https://www.cnblogs.com/chongzi1990/p/8694883.html
        # 去除 \ufeff. 直接指定  encoding='UTF-8-sig'
        with open(txt_path, encoding='UTF-8-sig') as txt_file:
            txt_content = txt_file.read().replace("\n\n", "\r\n")

            # 下一句代码是用来处理 json.dump(_json, file_obj, indent=4, ensure_ascii=False) 生成的txt文本正文
            # txt_content_out = txt_content.replace("[", "").replace("]", "").replace("\"", "").replace("    ","").replace("\n","")

            txtlist = txt_content.split("\r\n")

            try:
                app = xlwings.App(visible=False, add_book=False)
                time.sleep(0.5)
                book = app.books.add()

                # book.sheets.add("Sheet2")  # 添加一个Sheet
                # book.sheets.__delitem__("Sheet1")  # 删除Sheet1

                # 这种一行一行的写入方式效率太低
                #for j in range(len(txtlist)):
                #    xlwings.Range((j + 1, 1)).value = txtlist[j]  # xlwing从1开始
                #    print("Excel文档" + sheet_name + ", 第" + str(j + 1) + "行写完了")

                # 试试xlwings的range.value直接赋值的方式，注意添加.options(ndim=2)，不然会横着写
                # 逐行读取，塞到list里，会不会太慢了？
                multiline_txt_values = []
                for j in range(len(txtlist)):
                    multiline_txt_values.append(txtlist[j])

                sheet = xlwings.Sheet(xlwings.sheets[0])
                sheet.range("a1:a" + str(len(txtlist))).options(ndim=2).value = txt_content


                book.save(excel_path)
                book.close()
                app.quit()  # 【坑】这里千万只能写quit，不能直接kill否则循环的下次就跑不动了
                time.sleep(1)  # 防止Excel压力太大，睡1秒
                print("呵呵呵")
            except:
                traceback.print_exc()
            finally:
                app.quit()
                app.kill()


def write_txt_to_xlsx(txt_path, excel_path):
    print("debug only")
    sheet_name = excel_path.split(".xlsx")[0].split("\\")[-1]  # 需要创建的Excel文档的Sheet1重命名为产品名
    with open(txt_path, encoding='UTF-8-sig') as txt_file:
        txt_content = txt_file.read().replace("\n\n", "\r\n")

        # 下一句代码是用来处理 json.dump(_json, file_obj, indent=4, ensure_ascii=False) 生成的txt文本正文
        # txt_content_out = txt_content.replace("[", "").replace("]", "").replace("\"", "").replace("    ","").replace("\n","")

        txtlist = txt_content.split("\r\n")

        try:
            app = xlwings.App(visible=False, add_book=False)
            #time.sleep(0.5)
            book = app.books.open(excel_path)

            # book.sheets.add("Sheet2")  # 添加一个Sheet
            # book.sheets.__delitem__("Sheet1")  # 删除Sheet1

            for j in range(len(txtlist)):
                xlwings.Range((j + 1, 1)).value = txtlist[j]  # xlwing从1开始
                print("Excel文档" + sheet_name + ", 第" + str(j + 1) + "行写完了")

            book.save(excel_path)
            book.close()
            app.quit()  # 【坑】这里千万只能写quit，不能直接kill否则循环的下次就跑不动了
            #time.sleep(1)  # 防止Excel压力太大，睡1秒
            print("呵呵呵")
        except:
            traceback.print_exc()
        finally:
            app.quit()
            app.kill()


def get_column_count():
    '''
    获取Excel有效列数
    :return: 列数
    '''
    num = 0
    for i in range(1,255):
        if xlwings.Range((1,i)).value is not None:
            #print(xlwings.Range((1,i)).value)
            num +=1
        else:
            break  # 否则直接跳出循环，减少循环次数，提高效率
    return num


def get_column_value_list(row=1):
    _col_value_list = []
    for i in range(1,255):
        _col_value= xlwings.Range((row,i)).value
        if _col_value is not None:
            _col_value_list.append(_col_value)
        else:
            break  # 否则直接跳出循环，减少循环次数，提高效率
    return _col_value_list


def get_row_count():
    '''
    获取Excel有效行数（效率很低）【废弃】
    :return: 行数
    '''
    num = 0
    for i in range(1,65535):
        if xlwings.Range((i,1)).value is not None:
            #print(xlwings.Range((i,1)).value)
            num +=1
        else:
            break  # 否则直接跳出循环，减少循环次数，提高效率
    return num


'''
def merg_excel(dir):
    if dir[-1:] != "\\":
        dir = dir + "\\"
    excel_file_list = os.listdir(dir)
    col_value_list = []  # 循环遍历所有文件，获取所有的列名

    for i in range(len(excel_file_list)):
        print("merging:  " + str(i))
        try:
            app = xlwings.App(visible=False, add_book=False)
            #time.sleep(1)
            book = app.books.open(dir + excel_file_list[i])

            _list = get_column_value_list(row=1)
            for each in _list:
                if each not in col_value_list:
                    col_value_list.append(each)
            book.close()
            app.quit()
        except:
            traceback.print_exc()
        finally:
            app.quit()
    print(col_value_list)
'''

def get_row0_values(dir):
    '''
    获取文件夹下，所有Excel文件的title，合并、去重、排序
    :param dir: 存放Excel的文件夹
    :return: title
    '''
    if dir[-1:] != "\\":
        dir = dir + "\\"
    excel_file_list = os.listdir(dir)
    row0_values = []  # 循环遍历所有文件，获取所有的列名
    #row_num_total = 0
    for i in range(len(excel_file_list)):
        print("analyze Excel document:  " + str(i))
        try:
            book = xlrd.open_workbook(dir + excel_file_list[i])
            sheet = book.sheet_by_index(0)
            row_0_values = sheet.row_values(0)  #标题栏，也就是row = 1 时的值
            #row_num = sheet.nrows  # 当前打开的Excel文档的行数
            #row_num_total = row_num_total + row_num  # 累加计算总行数，为后续写入做准备
            for j in range(len(row_0_values)):
                each = row_0_values[j]
                print(each)
                if each not in row0_values:
                    row0_values.append(each)
            book.release_resources()
        except:
            traceback.print_exc()
        finally:
            book.release_resources()
    #print(row0_values)

    move_item_to_end(row0_values, "价格")

    print(row0_values)  # debug only

    return row0_values

def merg_excel(dir, merge_destination_dir):
    '''
    # https://www.xlwings.org/
    # http://docs.xlwings.org/en/stable/datastructures.html#lists
    将指定文件夹下的Excel，合并写入到新建的Excel里。列求得并集。如果不存在的数据则空置
    :param dir: 存放Excel的文件夹
    :return: 返回一个Excel文档，里面是合并后的数据
    '''
    if dir[-1:] != "\\":
        dir = dir + "\\"
    if merge_destination_dir[-1:] != "\\":
        merge_destination_dir = merge_destination_dir + "\\"
    excel_file_list = os.listdir(dir)

    row_0_values_merge = get_row0_values(dir)  # 合并标题去重后的总标题
    max_col_merge = len(row_0_values_merge)  # 合并的Excel表格的列长度

    # 存放合并文件的地方, 写文件我们用openpyxl
    # https://www.cnblogs.com/kearney908/p/7083958.html

    app_merge = xlwings.App(visible=False, add_book=False)
    book_merge = app_merge.books.add()
    book_merge.save(merge_destination_dir + "dell_合并.xlsx")
    sheet_merge = xlwings.Sheet(xlwings.sheets[0])

    print("新建成功！！！")

    merge_row_start_position = 2  # 写入数据的初始行数，第二行。合并表中的初始值是2
    current_row_start_position = 2  # 当前Excel，读取数据的起始行，常数2

    # 1.先把标题列填充
    sheet_merge.range("A1", sheet_merge["%s%d" % (chr(97 + max_col_merge-1),1)]).value = row_0_values_merge
    book_merge.save()  # 写完标题数据后保存当前Excel

    for i in range(len(excel_file_list)):  # 循环的次数为文件个数
        print("writing:  " + str(i))
        try:
            #excel = openpyxl.load_workbook(dir + excel_file_list[i])  # 打开当前文档，准备写入目标文档
            #sheet = excel.get_sheet_by_name(excel.get_sheet_names()[0])  # 打开的第一个sheet
            #max_row = sheet.max_row
            #max_col = sheet.max_column
            app_current = xlwings.App(visible=False, add_book=False)
            book_current = app_current.books.open(dir + excel_file_list[i])
            sheet_current = xlwings.Sheet(xlwings.sheets[0])

            max_row_current = xlwings.Range("A1").expand("table").last_cell.row - 1  # 打开当前的Excel，将数据写入到合并的Excel，要去掉标题栏
            max_col_current = xlwings.Range("A1").expand("table").last_cell.column

            row_0_values_current = sheet_current.range("A1", sheet_current["%s%d" % (chr(97 + max_col_current-1),1)]).value # 当前Excel# 标题值


            index_current_in_merge_list = []  # 当前标题在总标题中对应的位置的list
            for x in range(len(row_0_values_current)):
                if row_0_values_merge.index(row_0_values_current[x]) >= 0:
                    index_current_in_merge_list.append(row_0_values_merge.index(row_0_values_current[x]))  # 获取当前标题在总标题中对应的位置
                else:
                    continue

            range_index_list, range_current_list = get_continuous_number(index_current_in_merge_list)  # 返回一个list，这里返回的是<class 'list'>: ['1:8', 11]

            # 2.再填充其他的
            #for m in range(merge_row_start_position, max_row_current + merge_row_start_position):  # Excel的行数起始值是position，postition初始值是1
            for n in range(len(range_index_list)):
                if isinstance(range_index_list[n], tuple):  # 如果是tuple  (97,105)，则拆分返回
                    (col_merge_start, col_merge_end) = range_index_list[n]
                    col_current_start = int(range_current_list[n][0])
                    col_current_end = int(range_current_list[n][1])
                else:  # 否则直接返回
                    col_merge_start = col_merge_end = range_index_list[n]  # 起始和结束 都赋值给同一个
                    col_current_start = col_current_end = range_current_list[n]

                col_delta = int(col_merge_end) - int(col_merge_start) + 1  # 用来计数，指的是A2:I502, I和A之间的数量

                # 计算总表中的cell位置
                col_merge_start = chr(int(col_merge_start))  # 97 转成a
                col_merge_end = chr(int(col_merge_end))
                cell_merge_start = '%s%d' % (col_merge_start, merge_row_start_position)
                cell_merge_end = '%s%d' % (col_merge_end, merge_row_start_position + max_row_current-1)

                # 计算当前表的cell位置，每次需要加上col_position_current
                col_current_start = chr(col_current_start)
                col_current_end = chr(col_current_end)
                cell_current_start = '%s%d' % (col_current_start, current_row_start_position)  # 2是起始位置
                cell_current_end = '%s%d' % (col_current_end, current_row_start_position + max_row_current-1)  #

                merge_range = cell_merge_start + ":" + cell_merge_end
                current_range = cell_current_start + ":" + cell_current_end

                # 将当前表current的范围的值赋值给总表merge表 options(ndim=2)表示列的方向向下
                sheet_merge.range(merge_range).value = sheet_current.range(current_range).options(ndim=2).value
                #sheet_merge.range(merge_range).autofit()  # 调整格式
                del col_merge_start, col_merge_end, col_current_start, col_current_end  # 循环完了删除变量
                del merge_range
                del current_range
            app_current.quit()
            del index_current_in_merge_list[:]
            del row_0_values_current[:]
            del range_index_list[:]
            del range_current_list[:]
            del x
            merge_row_start_position += max_row_current  # 将Excel的位置加上当前打开的Excel的最大row数量，更新位置，供下次下入调整位置。
            book_merge.save()  # 写完所有的数据后保存当前Excel
        except:
            traceback.print_exc()
        finally:
            pass

    sheet_merge.autofit()  #最后，将所有单元格autofit一下，调整单元格

    book_merge.close()  # 关闭
    app_merge.quit()  # 退出当前的Excel进程
    app_merge.kill()


# 尝试用ProcessPoolExecutor代替ThreadPoolExecutor，绕过Global Interpretor Lock的限制
def write_txt_to_xlsx_batch(_txtrootpath, _excelrootpath):
    # 判断传入的路径的两种方式。C:\Folder和C:\Folder统一处理成C:\Folder\
    if _txtrootpath[-1:] != "\\":
        txtrootpath = _txtrootpath + "\\"
    if _excelrootpath[-1:] != "\\":
        excelrootpath = _excelrootpath + "\\"

    _txt_file_list = os.listdir(_txtrootpath)
    _excel_file_list = os.listdir(_excelrootpath)

    file_length = len(os.listdir(_txtrootpath))
    # MAX_WORKERS = os.cpu_count() 依赖当前运行环境的多核
    _max_workers = min(4, file_length)
    with futures.ProcessPoolExecutor(max_workers=_max_workers) as executor:
        to_do = []
        count_future = 0  # 声明一个计数器，用来计算进程数，并方便打印
        # 用于创建future

        for i in range(file_length):
            if i == 8 : break  # 一旦发生异常，可根据已完成的数量重新调整循环起始点
            _excel_path = _excelrootpath + _excel_file_list[i]  # 构造excel的绝对路径
            _txt_path = _txtrootpath + _txt_file_list[i].split(".txt")[0] + ".txt"
            # submit方法排定可调用对象的执行时间后返回一个future，表示这个待执行的操作
            # submit第一个参数是 获取格式数据和价格的方法write_txt_to_xlsx
            # 第二个参数是 submit第一个参数(方法名)的参数的 一个子元素， 即每次循环遍历出来的一个
            _future = executor.submit(write_txt_to_xlsx, _txt_path, _excel_path)
            to_do.append(_future)
            print("++++++++++在执行第  " + str(count_future) + "  个循环++++++++++")
            #time.sleep(1)
            count_future += 1

        results = []
        # 用于获取future的结果
        # as_completed 接收一个future列表，返回值是一个迭代器，在运行结束后产出future

        for future in futures.as_completed(to_do):
            _result = future.result()
            print("收到结果")
            results.append(_result)
    return len(results)

def main(write_txt_to_xlsx_batch):  # <10>
    t0 = time.time()
    count = write_txt_to_xlsx_batch(_txtrootpath, _excelrootpath)
    elapsed = time.time() - t0
    msg = '\n{} flags downloaded in {:.2f}s'
    print(msg.format(count, elapsed))




if __name__ == '__main__':
    dell_list = ['戴尔 Inspiron 灵越 14', '戴尔 Inspiron 灵越 15R', '戴尔 Inspiron 灵越 15 5000', '戴尔 Inspiron 1440', '戴尔 Inspiron 灵越 14R（7420）特别版', '戴尔 Inspiron 灵越 14 5000', '戴尔 XPS 11', '戴尔 6000', '戴尔 Inspiron 灵越 15（3521）', '戴尔 D630', '戴尔 Vostro 成就 5460', '戴尔 D610', '戴尔 XPS 13（9350）', '戴尔 Inspiron 灵越 M531R', '戴尔 Inspiron 灵越 15 3000（3558）', '戴尔 Inspiron 灵越 M431R', '戴尔 新Inspiron 灵越 15R', '戴尔 New Studio 思跃 17', '戴尔 Vostro 成就 V130', '戴尔 Vostro 成就 3300', '戴尔 1300', '戴尔 500', '戴尔 Latitude E5520', '戴尔 Vostro 成就 1440', '戴尔 A860', '戴尔 600m', '戴尔 Latitude E5400', '戴尔 D430', '戴尔 Adamo', '戴尔 D620', '戴尔 D800', '戴尔 Inspiron 灵越 14R（N4120）锋型版', '戴尔 Inspiron 灵越 15 7000游匣7000（7557）', '戴尔 Latitude E6330', '戴尔 1500', '戴尔 Inspiron 灵越 17R（5721）', '戴尔 Inspiron 灵越 14（3421）', '戴尔 Studio XPS 13', '戴尔 Vostro 成就 3460', '戴尔 Vostro 成就 1520', '戴尔 1427', '戴尔 Vostro 成就 14 5000（5480）', '戴尔 Inspiron 灵越 14 5000出彩版', '戴尔 Latitude E5530', '戴尔 Vostro 成就 2421', '戴尔 1450', '戴尔 Inspiron 灵越 15 3000', '戴尔 Vostro 成就 1320', '戴尔 710m', '戴尔 D530', '戴尔 XPS 12', '戴尔 Latitude XT3', '戴尔 Vostro 成就 2521', '戴尔 700m', '戴尔 Inspiron 灵越 17 5000（5748）', '戴尔 1318', '戴尔 Inspiron 灵越 M102z', '戴尔 Vostro 成就 1015', '戴尔 1000', '戴尔 Inspiron 灵越 15 7000游匣7000（7559）', '戴尔 Latitude 3330', '戴尔 Latitude E6500', '戴尔 1425', '戴尔 Inspiron 灵越 14 3000（AMD）', '戴尔 Latitude E4200', '戴尔 Inspiron 灵越 13z（5323）', '戴尔 Inspiron 灵越 14R（N4110）换壳版', '戴尔 D520', '戴尔 Latitude E6430', '戴尔 1710', '戴尔 Latitude E5430', '戴尔 X1', '戴尔 Vostro 成就 V13', '戴尔 1555', '戴尔 D600', '戴尔 Inspiron 灵越 13 7000', '戴尔 M1710', '戴尔 Inspiron 灵越 M411R', '戴尔 Vostro 成就 5470', '戴尔 Inspiron 灵越 13R', '戴尔 Latitude E6400', '戴尔 Latitude E6230', '戴尔 Latitude 14 5000（E5470）', '戴尔 Inspiron 灵越 M521R', '戴尔 Vostro 成就 3500', '戴尔 XPS 15（9550）', '戴尔 17', '戴尔 Inspiron 灵越 13z', '戴尔 100L', '戴尔 Inspiron 灵越 14（3437）', '戴尔 Inspiron 灵越 11 3000（3138）', '戴尔 Vostro 成就 15 3000（3559）', '戴尔 Inspiron 灵越 15 5000出彩版（5559）', '戴尔 5100', '戴尔 Inspiron 灵越 15（3537）', '戴尔 Latitude E6420', '戴尔 Inspiron 灵越 M301Z', '戴尔 燃7000', '戴尔 Inspiron 灵越 15 3000（AMD）', '戴尔 Vostro 成就 V131', '戴尔 XPS 12二合一笔记本（9250）', '戴尔 1520', '戴尔 Inspiron 灵越 15R（5537）', '戴尔 Inspiron 灵越 M511R', '戴尔 1100', '戴尔 Vostro 成就 1014', '戴尔 Vostro 成就 14 3000（3459）', '戴尔 Inspiron 灵越 M421R', '戴尔 Inspiron 灵越 13 7000多彩（7359）', '戴尔 Vostro 成就 3350', '戴尔 1720', '戴尔 Latitude 14 7000（E7470）', '戴尔 XPS 15', '戴尔 Vostro 成就 14 5000（5459）', '戴尔 D520n', '戴尔 Inspiron 灵越 14 5000出彩版AMD（5455）', '戴尔 Inspiron 灵越 14 5000金属版（5457）', '戴尔 Inspiron 灵越 17 7000', '戴尔 XPS 13', '戴尔 1536', '戴尔 Latitude 14 3000（3470）', '戴尔 D510', '戴尔 Inspiron 灵越 14R（5421）触控版', '戴尔 Inspiron 灵越 17 5000出彩版（5759）', '戴尔 Latitude 15 3000（3570）', '戴尔 Vostro 成就 2420', '戴尔 Inspiron 灵越 15R Turbo', '戴尔 Latitude 12 7000（E7270）', '戴尔 Studio 思跃 15z', '戴尔 Inspiron 灵越 14 5000（5442）', '戴尔 Latitude 15 5000（E5570）', '戴尔 M70', '戴尔 Inspiron 灵越 14z', '戴尔 Latitude 14 3000（3460）', '戴尔 Latitude E5420', '戴尔 XPS 15z', '戴尔 Latitude 12 5000（E5270）', '戴尔 Latitude 13（3350）', '戴尔 Vostro 成就 1450', '戴尔 A840', '戴尔 Latitude 13 7000（E7370）', '戴尔 Latitude 3440', '戴尔 Latitude E4310', '戴尔 510M', '戴尔 Vostro 成就 1720', '戴尔 新XPS 15', '戴尔 1200', '戴尔 1501', '戴尔 New Studio 思跃 14', '戴尔 Vostro 成就 3750', '戴尔 1150', '戴尔 Inspiron 灵越 14 3000', '戴尔 Vostro 成就 1088', '戴尔 Latitude E6530', '戴尔 Vostro 成就 15 3000', '戴尔 Inspiron 灵越 15R（5521）', '戴尔 Vostro 成就 2520', '戴尔 D505', '戴尔 Latitude 13', '戴尔 Latitude 13 7000', '戴尔 M1210', '戴尔 X300', '戴尔 Inspiron 灵越 15', '戴尔 8500', '戴尔 Inspiron 灵越 15 5000金属版（5557）', '戴尔 M1530', '戴尔 640m', '戴尔 Inspiron 灵越 11 3000', '戴尔 Inspiron 灵越 15 7000（7537）', '戴尔 2200', '戴尔 XPS 16', '戴尔 1435', '戴尔 Inspiron 灵越 M101z', '戴尔 N4010', '戴尔 6400', '戴尔 M1330', '戴尔 全新Inspiron 灵越 15 7000', '戴尔 Inspiron 灵越 14 3000（3458）', '戴尔 Inspiron 灵越 14R（5421）', '戴尔 Inspiron 灵越 11z', '戴尔 Inspiron 灵越 14R（5437）', '戴尔 M1730', '戴尔 630m', '戴尔 Inspiron 灵越 15 5000（AMD）', '戴尔 D410', '戴尔 Vostro 成就 1220', '戴尔 Inspiron 灵越 M501R', '戴尔 Vostro 成就 3560', '戴尔 1310', '戴尔 1510', '戴尔 Vostro 成就 5560', '戴尔 Inspiron 灵越 14R Turbo', '戴尔 Latitude E4300', '戴尔 D420', '戴尔 Inspiron 灵越 14 5000（AMD）', '戴尔 Inspiron 灵越 14 7000（游匣7000）', '戴尔 D500', '戴尔 XPS 14z', '戴尔 Vostro 成就 3550', '戴尔 Latitude XT2', '戴尔 XPS 14', '戴尔 Latitude E6520', '戴尔 8600', '戴尔 D820', '戴尔 Inspiron 灵越15 5000(5565)', '戴尔 Vostro 成就 14 3000', '戴尔 Vostro 成就 3700', '戴尔 1400', '戴尔 Vostro 成就 3450', '戴尔 9300', '戴尔 Ins15-7560', '戴尔 Inspiron 灵越 15 5000（5542）', '戴尔 Inspiron 灵越 14 7000', '戴尔 1410', '戴尔 Studio 思跃 15', '戴尔 5000系列出彩版', '戴尔 Ins15-7566', '戴尔 XPS 13-9360', '戴尔 XPS13D-9343', '戴尔 灵越15 3000(3568)新飞匣', '戴尔 Ins14-7460', '戴尔 Inspiron 灵越 17R Turbo', '戴尔 Inspiron灵越14 5000(5525)', '戴尔 Latitude E6220', '戴尔 1420', '戴尔 Ins15-5567', '戴尔 Inspiron 灵越 14 5000（5439）', '戴尔 Inspiron 灵越 15R（7520）特别版', '戴尔 XPS 17', '戴尔 Latitude E6430s', '戴尔 XPS 15-9550', '戴尔 1526', '戴尔 1525', '戴尔 Ins14-7466', '戴尔 燃7000 II（7472）', '戴尔 131L', '戴尔 Vostro 成就 15 5000系列 5568', '戴尔 C640', '戴尔 D531', '戴尔 Inspiron 灵越 15 7000（7577）', '戴尔 Inspiron灵越14 7000(7437)', '戴尔 D810', '戴尔 Ins15PD-3948', '戴尔 Inspiron 灵越 13', '戴尔 XPS L502X', '戴尔 XPS15-9560', '戴尔 Inspiron 15-7567', '戴尔 N4030', '戴尔 Vostro 成就 14 5000系列 5468', '戴尔 Vostro 成就 3400', '戴尔 E7450', '戴尔 Vostro成就14 3000(3458)', '戴尔 Inspiron 灵越 15 5000（5570）', '戴尔 Latitude E6410', '戴尔 M5455D-1828', '戴尔 M5545D', '戴尔 M5555D-2828', '戴尔 XPS 13-9350-D2508', '戴尔 500M', '戴尔 9400', '戴尔 Ins15PD-2748', '戴尔 Inspiron 15-5577', '戴尔 Inspiron 灵越 14V', '戴尔 Inspiron灵越 15 5576(游匣Speed) 系列', '戴尔 Latitude 14 5480', '戴尔 Vostro 成就 15 7000（7570）', '戴尔 Ins 11-3162', '戴尔 Ins 13MF', '戴尔 Ins14VR', '戴尔 Ins15ED', '戴尔 Ins15PD-1548', '戴尔 Inspiron 15-7570', '戴尔 Inspiron 灵越15（I1564D-138）', '戴尔 Latitude E5440', '戴尔 Latitude E6320', '戴尔 XPS 13-9350-D2708', '戴尔 XPS 13-9350-D3608', '戴尔 XPS15D', '戴尔 Ins14SD-4528', '戴尔 Ins14UD-3528', '戴尔 Ins15CD-4528', '戴尔 Ins17-5767', '戴尔 Inspiron 15-3567', '戴尔 Inspiron 灵越 14 5000出彩版（5459）', '戴尔 Inspiron 灵越 17 7000系列 魔方', '戴尔 Vostro 成就 15 3000系列 3562', '戴尔 XPS 12-9250', '戴尔 XPS 13-9350-D1608', '戴尔 XPS 13-9350-D2608', '戴尔 Ins14PD-2748', '戴尔 Ins14PD-3548', '戴尔 Ins14SD-4328', '戴尔 Ins15MD', '戴尔 Ins15PD-1748', '戴尔 Ins15PD-2749', '戴尔 Ins15UD-2528', '戴尔 Ins15UD-3528', '戴尔 Ins15UD-3548', '戴尔 Inspiron灵越 11 3157', '戴尔 Inspiron灵越14 3000(3459)', '戴尔 Latitude 3550', '戴尔 M5455D-2828', '戴尔 M5555D-1828', '戴尔 XPS 13-9350-D2808', '戴尔 XPS 13-9350-D3808', '戴尔 XPS 15 系列（9560）', '戴尔 Ins 15MF', '戴尔 Ins11BD-1208', '戴尔 Ins13BD-1508', '戴尔 Ins14ED', '戴尔 Ins14MD', '戴尔 Ins14UD-3748', '戴尔 Ins14UD-4728', '戴尔 Ins15AD', '戴尔 Ins15PD-1848', '戴尔 Ins15PD-2548', '戴尔 Ins15PD-2648', '戴尔 Ins15PD-3848', '戴尔 Ins15UD-3508', '戴尔 Inspiron 13-7370', '戴尔 Inspiron灵越 13 7000(7359)', '戴尔 Latitude 14 5000（5490）', '戴尔 Latitude 5175', '戴尔 Latitude E5410', '戴尔 XPS 13-9350-D1609', '戴尔 XPS 13-9350-D1708', '戴尔 XPS 13-9350-D4505', '戴尔 XPS 13D-9343-5508', '戴尔 XPS-13-9333D', '戴尔 灵越13 7368', '戴尔 Chromebook 13', '戴尔 E5250', '戴尔 E5550', '戴尔 E6440', '戴尔 E6540', '戴尔 E7250', '戴尔 E7470', '戴尔 G3 15游戏本（3579）', '戴尔 G3 17游戏本（3779）', '戴尔 G5', '戴尔 G7', '戴尔 Ins11BD-1308', '戴尔 Ins11WD', '戴尔 Ins13BD-1708', '戴尔 Ins13WD', '戴尔 Ins14PD-2548', '戴尔 Ins14PD-2848', '戴尔 Ins14PD-2948', '戴尔 Ins14PD-3748', '戴尔 Ins14PD-4548', '戴尔 Ins14PD-4748', '戴尔 Ins14SD-1116', '戴尔 Ins14SD-1328', '戴尔 Ins14SD-1528', '戴尔 Ins14SD-3328', '戴尔 Ins14SD-3528', '戴尔 Ins14UD-1108', '戴尔 Ins14UD-1328', '戴尔 Ins14UD-1528', '戴尔 Ins14UD-1548', '戴尔 Ins14UD-1628', '戴尔 Ins14UD-1748', '戴尔 Ins14UD-2108', '戴尔 Ins14UD-2328', '戴尔 Ins14UD-2528', '戴尔 Ins14UD-3108', '戴尔 Ins14UD-3548', '戴尔 Ins14UD-4108', '戴尔 Ins14UD-4328', '戴尔 Ins14UD-4528', '戴尔 Ins14UD-4548', '戴尔 Ins14UD-5525', '戴尔 Ins14UD-5625', '戴尔 Ins14UD-5725', '戴尔 Ins15BD', '戴尔 Ins15CD-2328', '戴尔 Ins15PD-4848', '戴尔 Ins15SD-1308', '戴尔 Ins15SD-1328', '戴尔 Ins15SD-1528', '戴尔 Ins15SD-3328', '戴尔 Ins15SD-3528', '戴尔 Ins15SD-4328', '戴尔 Ins15SD-4528', '戴尔 Ins15UD-1108', '戴尔 Ins15UD-1328', '戴尔 Ins15UD-1528', '戴尔 Ins15UD-1548', '戴尔 Ins15UD-1748', '戴尔 Ins15UD-2108', '戴尔 Ins15UD-2328', '戴尔 Ins15UD-3328', '戴尔 Ins15UD-3628', '戴尔 Ins15UD-3748', '戴尔 Ins15UD-4328', '戴尔 Ins15UD-4748', '戴尔 Ins15UD-5748', '戴尔 Ins15WD-1508', '戴尔 Ins15WD-1608', '戴尔 Ins15WD-1708', '戴尔 Ins17HD', '戴尔 Ins17UD', '戴尔 Inspiron 11-3168', '戴尔 Inspiron 11-3179', '戴尔 Inspiron 13-5378', '戴尔 Inspiron 13-5379', '戴尔 Inspiron 13-7373', '戴尔 Inspiron 14-3462', '戴尔 Inspiron 14-3465', '戴尔 Inspiron 14-3467', '戴尔 Inspiron 14-3468', '戴尔 Inspiron 14-3476', '戴尔 Inspiron 14-5468', '戴尔 Inspiron 14-7467', '戴尔 Inspiron 15-3552', '戴尔 Inspiron 15-3559', '戴尔 Inspiron 15-3575', '戴尔 Inspiron 15-3576', '戴尔 Inspiron 15-5575', '戴尔 Inspiron 15-5578', '戴尔 Inspiron 15-7572', '戴尔 Inspiron 15-7579', '戴尔 Inspiron 15-7779', '戴尔 Inspiron 灵越 12 5000（5280）', '戴尔 Inspiron 灵越 13 5000（5370）', '戴尔 Inspiron 灵越 13 7000(7347)', '戴尔 Inspiron 灵越 14 5000(5482）', '戴尔 Inspiron灵越13 5000(7378)', '戴尔 Latitude 11 5000系列', '戴尔 Latitude 12 5000系列', '戴尔 Latitude 12 7000 7275', '戴尔 Latitude 12 7000系列 7280', '戴尔 Latitude 12 7000系列 7285', '戴尔 Latitude 13 3000系列', '戴尔 Latitude 13 7350', '戴尔 Latitude 14 3000系列', '戴尔 Latitude 14 3000（3490）', '戴尔 Latitude 14 7000 7480', '戴尔 Latitude 15 3000系列 3560', '戴尔 Latitude 15 3000系列 3580', '戴尔 Latitude 15 5000系列 5580', '戴尔 Latitude 3379', '戴尔 Latitude 3380', '戴尔 Latitude 3450', '戴尔 Latitude 3480', '戴尔 Latitude 3560', '戴尔 Latitude 5280', '戴尔 Latitude 5289', '戴尔 Latitude 7380', '戴尔 Latitude 7389', '戴尔 Latitude E6510', '戴尔 Latitude E7240', '戴尔 M5455D-3208', '戴尔 M5455D-3828', '戴尔 M5455D-4208', '戴尔 M5455D-4828', '戴尔 M5455D-5205', '戴尔 M5555D-2928', '戴尔 Precision3510', '戴尔 Precision7520', '戴尔 Precision7720', '戴尔 Vostro 15-3572', '戴尔 Vostro 3561', '戴尔 Vostro 成就 13 5000（5370）', '戴尔 Vostro 成就 14 3000系列 3468', '戴尔 Vostro 成就 14 3000系列 3478', '戴尔 Vostro 成就 14 5000系列 5471', '戴尔 Vostro 成就 15 3000系列 3546', '戴尔 Vostro 成就 15 3000系列 3568', '戴尔 Vostro 成就 15 3000系列 3578', '戴尔 Vostro成就15 3000(3558)', '戴尔 XPS 13-9350-D1508', '戴尔 XPS 13-9350-D1808', '戴尔 XPS 13-9350-D3708', '戴尔 XPS 13-9365', '戴尔 XPS 13-9370', '戴尔 XPS L701X', '戴尔 灵越13-5368', '戴尔 M20', '戴尔 Vostro 成就 3360', '戴尔 Latitude E6430 ATG']
    #print(str(dell_list))

    dell_current = []
    filenames = os.listdir("F:\\电脑报价\\youdemai\\dell\\500_excel\\")
    for filename in filenames:
        filename = filename.split(".xlsx")[0]
        dell_current.append(filename)
    print(dell_current)
    # 求差集
    ret_list =[]
    for dell in dell_current:
        if dell in dell_list:
            ret_list.append(dell_list.index(dell))

    print(str(ret_list))
    # index = [306, 354, 357, 358, 359, 362, 366, 311, 368, 369, 370, 312, 372, 373]


    source_dir = u"F:\\电脑报价\\youdemai\\dell\\500\\"
    destination_dir = u"F:\\电脑报价\\youdemai\\dell\\500_excel_empty\\"
    #create_file_from_file(source_dir, destination_dir)

    # 方式1：直接运行
    write_txts_to_xlsxs(u"F:\\电脑报价\\youdemai\\dell\\500\\", u"F:\\电脑报价\\youdemai\\dell\\500_excel\\")

    # 方式2：多进程
    _txtrootpath = u"F:\\电脑报价\\youdemai\\dell\\500_backup\\"
    _excelrootpath = u"F:\\电脑报价\\youdemai\\dell\\500_backup_excel\\"
    #main(write_txt_to_xlsx_batch)
    #time.sleep(10)
    #merg_excel("F:\\PycharmWorkSpace\\Learn\\1000_excel_合并测试")
    merge_destination_dir = "F:\\电脑报价\\youdemai\\dell\\"
    merg_excel("F:\\电脑报价\\youdemai\\dell\\500_excel_backup\\", merge_destination_dir)




